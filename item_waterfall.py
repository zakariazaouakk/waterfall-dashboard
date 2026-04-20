import io
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from utils import (
    load_excel_data, blank_pre_snapshot_weeks, apply_excel_formatting,
    compute_variation, empty_df, year_week,
    VAR_COLS,
)

def generate_item_waterfall(files_bytes_list):
    """Generate item-aggregated waterfall Excel file (one row per Item Number)."""
    excel_data, snapshot_weeks, all_weeks_set = load_excel_data(files_bytes_list)
    all_weeks = sorted(all_weeks_set, key=lambda x: (int(x.split("-")[1]), int(x[1:].split("-")[0])))

    # ── Unique item numbers across all files ──────────────────────────────────
    unique_items_df = pd.concat([
        pd.concat([
            excel_dict.get("Firm",     empty_df())[["Item Number"]],
            excel_dict.get("Forecast", empty_df())[["Item Number"]],
        ])
        for excel_dict in excel_data
    ]).drop_duplicates().sort_values("Item Number").reset_index(drop=True)

    # ── Build waterfall rows ──────────────────────────────────────────────────
    item_rows            = []
    item_row_file_indices = []

    for _, item_row in unique_items_df.iterrows():
        item_num = item_row["Item Number"]

        for file_idx, excel_dict in enumerate(excel_data):
            row_dict = {
                "Item Number":  item_num,
                "SnapshotWeek": f"CW{snapshot_weeks[file_idx]:02d}",
                **{w: 0 for w in all_weeks},
            }

            df_firm     = excel_dict.get("Firm",     empty_df())
            df_forecast = excel_dict.get("Forecast", empty_df())

            firm_rows     = df_firm    [df_firm    ["Item Number"] == item_num].copy()
            forecast_rows = df_forecast[df_forecast["Item Number"] == item_num].copy()

            # (Sales Order, Customer Item, DateStr) keys that have firm data
            firm_keys = set(zip(firm_rows["Sales Order"],
                                firm_rows["Customer Item"],
                                firm_rows["DateStr"]))

            # Aggregate firm quantities
            for _, r in (firm_rows
                         .groupby(["Sales Order", "Customer Item", "DateStr"], as_index=False)["Quantity"]
                         .sum()
                         .iterrows()):
                row_dict[year_week(pd.to_datetime(r["DateStr"]))] += r["Quantity"]

            # Add forecast only where no firm exists
            for _, r in forecast_rows.iterrows():
                if (r["Sales Order"], r["Customer Item"], r["DateStr"]) not in firm_keys:
                    row_dict[year_week(pd.to_datetime(r["DateStr"]))] += r["Quantity"]

            item_rows.append(row_dict)
            item_row_file_indices.append(file_idx)

        # Separator row
        item_rows.append({col: "" for col in ["Item Number", "SnapshotWeek"] + all_weeks})
        item_row_file_indices.append(None)

    # Remove trailing separator
    if item_rows:
        item_rows.pop()
        item_row_file_indices.pop()

    waterfall_items = pd.DataFrame(item_rows)

    blank_pre_snapshot_weeks(waterfall_items, item_row_file_indices, snapshot_weeks, all_weeks)

    # ── Variation columns ─────────────────────────────────────────────────────
    var_w1  = compute_variation(waterfall_items, item_row_file_indices, all_weeks, lookback=1)
    var_w2  = compute_variation(waterfall_items, item_row_file_indices, all_weeks, lookback=2)
    var_w4  = compute_variation(waterfall_items, item_row_file_indices, all_weeks, lookback=4)
    var_w13 = compute_variation(waterfall_items, item_row_file_indices, all_weeks, lookback=13)

    waterfall_items["W-1"]  = var_w1
    waterfall_items["W-2"]  = var_w2
    waterfall_items["W-4"]  = var_w4
    waterfall_items["W-13"] = var_w13

    item_cols_order = ["Item Number", "SnapshotWeek", "W-1", "W-2", "W-4", "W-13"] + all_weeks
    waterfall_items = waterfall_items[item_cols_order]

    # ── Write to Excel ────────────────────────────────────────────────────────
    output_buffer = io.BytesIO()
    waterfall_items.to_excel(output_buffer, index=False)
    output_buffer.seek(0)

    wb = load_workbook(output_buffer)
    ws = wb.active

    header          = [cell.value for cell in ws[1]]
    col_name_to_idx = {name: idx + 1 for idx, name in enumerate(header)}

    # Group index for alternating row colours
    group_index   = []
    current_group = -1
    current_key   = None
    for row_idx, file_idx in enumerate(item_row_file_indices):
        if file_idx is None:
            group_index.append(None)
        else:
            key = waterfall_items.at[row_idx, "Item Number"]
            if key != current_key:
                current_key = key
                current_group += 1
            group_index.append(current_group)

    apply_excel_formatting(
        ws              = ws,
        header          = header,
        col_name_to_idx = col_name_to_idx,
        row_file_indices= item_row_file_indices,
        group_index     = group_index,
        id_cols         = {"Item Number"},
        snapshot_weeks  = snapshot_weeks,
        all_weeks       = all_weeks,
        var_col_data    = {"W-1": var_w1, "W-2": var_w2, "W-4": var_w4, "W-13": var_w13},
    )

    col_widths = {
        "Item Number":  14, "SnapshotWeek": 12,
        "W-1": 9, "W-2": 9, "W-4": 9, "W-13": 9,
    }
    for col_idx, col_name in enumerate(header, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 11)

    ws.freeze_panes          = "C2"
    ws.row_dimensions[1].height = 22

    final_buffer = io.BytesIO()
    wb.save(final_buffer)
    final_buffer.seek(0)
    return final_buffer
