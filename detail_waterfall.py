import io
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from utils import (
    blank_pre_snapshot_weeks, apply_excel_formatting,
    compute_variation, empty_df, year_week,
)


def generate_detail_waterfall(pre_loaded):
    """
    Generate detail waterfall Excel (one row per Sales Order / Item / Customer Item).
    pre_loaded = (excel_data, snapshot_weeks, all_weeks_set) from agent.py
    """
    excel_data, snapshot_weeks, all_weeks_set = pre_loaded

    all_weeks = sorted(all_weeks_set,
                       key=lambda x: (int(x.split("-")[1]), int(x[1:].split("-")[0])))

    # ── Unique (Sales Order, Item Number, Customer Item) combinations ─────────
    unique_combinations = pd.concat([
        pd.concat([
            excel_dict.get("Firm",     empty_df())[["Sales Order", "Item Number", "Customer Item"]],
            excel_dict.get("Forecast", empty_df())[["Sales Order", "Item Number", "Customer Item"]],
        ])
        for excel_dict in excel_data
    ]).drop_duplicates()

    # ── Build waterfall rows ──────────────────────────────────────────────────
    waterfall_rows   = []
    row_file_indices = []

    for _, item in unique_combinations.iterrows():
        for file_idx, excel_dict in enumerate(excel_data):
            row_dict = {
                "Sales Order":   item["Sales Order"],
                "Item Number":   item["Item Number"],
                "Customer Item": item["Customer Item"],
                "SnapshotWeek":  f"CW{snapshot_weeks[file_idx]:02d}",
                **{w: 0 for w in all_weeks},
            }

            df_firm     = excel_dict.get("Firm",     empty_df())
            df_forecast = excel_dict.get("Forecast", empty_df())

            firm_mask = (
                (df_firm["Sales Order"]   == item["Sales Order"]) &
                (df_firm["Item Number"]   == item["Item Number"]) &
                (df_firm["Customer Item"] == item["Customer Item"])
            )
            firm_rows = df_firm[firm_mask].groupby("DateStr", as_index=False)["Quantity"].sum()
            firm_rows["YearWeek"] = pd.to_datetime(firm_rows["DateStr"]).apply(year_week)

            fore_mask = (
                (df_forecast["Sales Order"]   == item["Sales Order"]) &
                (df_forecast["Item Number"]   == item["Item Number"]) &
                (df_forecast["Customer Item"] == item["Customer Item"])
            )
            forecast_rows = df_forecast[fore_mask].groupby("DateStr", as_index=False)["Quantity"].sum()
            forecast_rows["YearWeek"] = pd.to_datetime(forecast_rows["DateStr"]).apply(year_week)

            firm_by_date = firm_rows.set_index("DateStr")["Quantity"].to_dict()

            for date_str, qty in firm_by_date.items():
                wk = year_week(pd.to_datetime(date_str))
                if wk in row_dict:
                    row_dict[wk] += qty

            for _, r in forecast_rows.iterrows():
                if r["DateStr"] not in firm_by_date:
                    wk = r["YearWeek"]
                    if wk in row_dict:
                        row_dict[wk] += r["Quantity"]

            waterfall_rows.append(row_dict)
            row_file_indices.append(file_idx)

        # Separator row
        waterfall_rows.append(
            {col: "" for col in
             ["Sales Order", "Item Number", "Customer Item", "SnapshotWeek"] + all_weeks}
        )
        row_file_indices.append(None)

    # Remove trailing separator
    if waterfall_rows:
        waterfall_rows.pop()
        row_file_indices.pop()

    waterfall = pd.DataFrame(waterfall_rows)
    blank_pre_snapshot_weeks(waterfall, row_file_indices, snapshot_weeks, all_weeks)

    # ── Variation columns ─────────────────────────────────────────────────────
    var_w1  = compute_variation(waterfall, row_file_indices, all_weeks, lookback=1)
    var_w2  = compute_variation(waterfall, row_file_indices, all_weeks, lookback=2)
    var_w4  = compute_variation(waterfall, row_file_indices, all_weeks, lookback=4)
    var_w13 = compute_variation(waterfall, row_file_indices, all_weeks, lookback=13)

    waterfall["W-1"]  = var_w1
    waterfall["W-2"]  = var_w2
    waterfall["W-4"]  = var_w4
    waterfall["W-13"] = var_w13

    cols_order = ["Sales Order", "Item Number", "Customer Item",
                  "SnapshotWeek", "W-1", "W-2", "W-4", "W-13"] + all_weeks
    waterfall = waterfall[cols_order]

    # ── Write to Excel ────────────────────────────────────────────────────────
    output_buffer = io.BytesIO()
    waterfall.to_excel(output_buffer, index=False)
    output_buffer.seek(0)

    wb = load_workbook(output_buffer)
    ws = wb.active

    header          = [cell.value for cell in ws[1]]
    col_name_to_idx = {name: idx + 1 for idx, name in enumerate(header)}

    group_index   = []
    current_group = -1
    current_key   = None
    for row_idx, file_idx in enumerate(row_file_indices):
        if file_idx is None:
            group_index.append(None)
        else:
            key = (waterfall.at[row_idx, "Sales Order"],
                   waterfall.at[row_idx, "Item Number"])
            if key != current_key:
                current_key = key
                current_group += 1
            group_index.append(current_group)

    apply_excel_formatting(
        ws               = ws,
        header           = header,
        col_name_to_idx  = col_name_to_idx,
        row_file_indices = row_file_indices,
        group_index      = group_index,
        id_cols          = {"Sales Order", "Item Number", "Customer Item"},
        snapshot_weeks   = snapshot_weeks,
        all_weeks        = all_weeks,
        var_col_data     = {"W-1": var_w1, "W-2": var_w2,
                            "W-4": var_w4, "W-13": var_w13},
    )

    col_widths = {
        "Sales Order": 14, "Item Number": 14, "Customer Item": 18,
        "SnapshotWeek": 12, "W-1": 9, "W-2": 9, "W-4": 9, "W-13": 9,
    }
    for col_idx, col_name in enumerate(header, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = \
            col_widths.get(col_name, 11)

    ws.freeze_panes             = "E2"
    ws.row_dimensions[1].height = 22

    final_buffer = io.BytesIO()
    wb.save(final_buffer)
    final_buffer.seek(0)
    return final_buffer
