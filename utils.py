import re
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment

# ── Sheet & column config ─────────────────────────────────────────────────────
VALID_SHEETS     = ["Deljit QAD extraction", "Delfor QAD extraction"]
REQUIRED_COLUMNS = ["Sales Order", "Item Number", "Customer Item", "Date", "Quantity"]

# ── Colour palette ────────────────────────────────────────────────────────────
HEADER_BG       = "1F4E79"
HEADER_FONT     = "FFFFFF"
VAR_COL_BG      = "BDD7EE"
VAR_COL_BG_ALT  = "9DC3E6"
ID_COL_BG       = "DEEAF1"
ID_COL_BG_ALT   = "B8CCE4"
WEEK_COL_BG     = "EBF3FB"
WEEK_COL_BG_ALT = "D6E4F0"
SEP_BG          = "F2F2F2"
YELLOW          = "FFFF00"
RED             = "FF0000"

# ── Pre-built style objects ───────────────────────────────────────────────────
def solid(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

header_fill  = solid(HEADER_BG)
header_font  = Font(name="Arial", bold=True, color=HEADER_FONT, size=10)
yellow_fill  = solid(YELLOW)
red_fill     = solid(RED)
white_font   = Font(name="Arial", color="FFFFFF", bold=True, size=10)
sep_fill     = solid(SEP_BG)
center_align = Alignment(horizontal="center", vertical="center")
left_align   = Alignment(horizontal="left",   vertical="center")

VAR_COLS      = {"W-1", "W-2", "W-4", "W-13"}
RED_THRESHOLD = {"W-1": 0.20, "W-2": 0.20, "W-4": 0.20, "W-13": 0.20}

# ── Parsing helpers ───────────────────────────────────────────────────────────
def extract_week_from_filename(filename: str) -> int:
    match = re.search(r"CW-?(\d+)", filename, re.IGNORECASE)
    return int(match.group(1)) if match else 0

def clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = df.columns.str.strip().str.replace("\n", "").str.replace("\r", "")
    return df

def clean_sheet(df: pd.DataFrame) -> pd.DataFrame:
    df = clean_columns(df)
    df = df[REQUIRED_COLUMNS].copy()
    for col in ["Sales Order", "Item Number", "Quantity"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df["Date"] = pd.to_datetime(df["Date"], dayfirst=True, errors="coerce")
    df = df.dropna(subset=["Sales Order", "Item Number", "Quantity", "Date"])
    return df

def year_week(date) -> str:
    iso = date.isocalendar()
    return f"W{iso[1]}-{iso[0]}"

def cw_str_to_int(cw) -> int:
    """Convert 'CW05' or 'CW5' → 5."""
    match = re.search(r"\d+", str(cw))
    return int(match.group()) if match else 0

# ── Variation computation ─────────────────────────────────────────────────────
def compute_variation(waterfall: pd.DataFrame, row_file_indices: list,
                      all_weeks: list, lookback: int) -> list:
    variation_col = []
    for row_idx, file_idx in enumerate(row_file_indices):
        if file_idx is None or file_idx < lookback:
            variation_col.append("")
            continue

        prev_row_idx  = row_idx - lookback
        snapshot_week = all_weeks[file_idx]
        curr_val      = waterfall.at[row_idx,      snapshot_week]
        prev_val      = waterfall.at[prev_row_idx, snapshot_week]

        try:
            curr_qty = float(curr_val)
            prev_qty = float(prev_val)
        except (ValueError, TypeError):
            variation_col.append("")
            continue

        if prev_qty == 0:
            variation_col.append("")
        else:
            variation_col.append((curr_qty - prev_qty) / prev_qty)

    return variation_col

# ── Empty DataFrame scaffold ──────────────────────────────────────────────────
_EMPTY_DF_COLS = ["Sales Order", "Item Number", "Customer Item",
                  "Date", "Quantity", "YearWeek", "DateStr", "SheetType"]

def empty_df() -> pd.DataFrame:
    return pd.DataFrame(columns=_EMPTY_DF_COLS)

# ── File loading ──────────────────────────────────────────────────────────────
def load_excel_data(files_bytes_list: list) -> tuple[list, list, set]:
    """
    Parse every uploaded CW source file into a list of {Firm/Forecast: DataFrame} dicts.

    Returns:
        excel_data     – list of dicts, one per file
        snapshot_weeks – list of CW integers, one per file
        all_weeks_set  – set of 'WXX-YYYY' strings seen across all files
    """
    excel_data    = []
    snapshot_weeks = []
    all_weeks_set  = set()

    for file_bytes, file_name in files_bytes_list:
        week_num = extract_week_from_filename(file_name)
        snapshot_weeks.append(week_num)

        xl         = pd.ExcelFile(file_bytes)
        excel_dict = {}
        for sheet_name in xl.sheet_names:
            if sheet_name in VALID_SHEETS:
                df      = xl.parse(sheet_name)
                df_type = "Firm" if "Deljit" in sheet_name else "Forecast"
                df      = clean_sheet(df)
                df["YearWeek"]  = df["Date"].apply(year_week)
                df["DateStr"]   = df["Date"].dt.strftime("%Y-%m-%d")
                df["SheetType"] = df_type
                excel_dict[df_type] = df
                all_weeks_set.update(df["YearWeek"].unique())
        excel_data.append(excel_dict)

    return excel_data, snapshot_weeks, all_weeks_set


def read_waterfall_snapshots(waterfall_bytes, waterfall_type: str) -> tuple[list, list, set]:
    """
    Reconstruct (excel_data, snapshot_weeks, all_weeks_set) from a previously
    generated waterfall Excel so it can be merged with new CW files.

    waterfall_type: "detail" → expects Sales Order, Item Number, Customer Item
                    "item"   → expects Item Number only
    """
    df = pd.read_excel(waterfall_bytes)

    # Identify week columns (pattern: W<n>-<yyyy>)
    week_col_pattern = re.compile(r"^W\d+-\d{4}$")
    week_cols     = [c for c in df.columns if week_col_pattern.match(str(c))]
    all_weeks_set = set(week_cols)

    id_cols = (["Sales Order", "Item Number", "Customer Item"]
               if waterfall_type == "detail" else ["Item Number"])

    # Drop separator rows
    df = df.dropna(subset=id_cols, how="all")
    df = df[df[id_cols[0]].astype(str).str.strip() != ""]

    # Each unique SnapshotWeek = one historical snapshot, sorted chronologically
    snapshot_cw_values = sorted(df["SnapshotWeek"].dropna().unique(), key=cw_str_to_int)

    excel_data     = []
    snapshot_weeks = []

    for cw_str in snapshot_cw_values:
        cw_int      = cw_str_to_int(cw_str)
        snapshot_df = df[df["SnapshotWeek"] == cw_str].copy()
        snapshot_weeks.append(cw_int)

        # Rebuild long-format rows from the week quantity columns.
        # Everything is treated as "Firm" — the original split doesn't matter
        # once the waterfall has been generated.
        rows = []
        for _, row in snapshot_df.iterrows():
            for wk in week_cols:
                val = row.get(wk, "")
                try:
                    qty = float(val)
                except (ValueError, TypeError):
                    continue
                if qty == 0:
                    continue

                # Reconstruct a representative date (Monday of that ISO week)
                try:
                    wk_num  = int(wk.split("-")[0][1:])
                    wk_year = int(wk.split("-")[1])
                    date    = pd.Timestamp.fromisocalendar(wk_year, wk_num, 1)
                except Exception:
                    continue

                rec = {
                    "Quantity":  qty,
                    "Date":      date,
                    "YearWeek":  wk,
                    "DateStr":   date.strftime("%Y-%m-%d"),
                    "SheetType": "Firm",
                }
                for col in id_cols:
                    rec[col] = row.get(col)
                # Pad any missing scaffold columns with None
                for col in _EMPTY_DF_COLS:
                    rec.setdefault(col, None)
                rows.append(rec)

        rebuilt    = pd.DataFrame(rows) if rows else empty_df()
        excel_data.append({"Firm": rebuilt})

    return excel_data, snapshot_weeks, all_weeks_set


def merge_excel_data(
    prev_data: tuple[list, list, set],
    new_data:  tuple[list, list, set],
) -> tuple[list, list, set]:
    """
    Merge historical snapshot data with newly uploaded CW file data.

    Raises ValueError listing any duplicate CW weeks so app.py can surface
    the error to the user and stop.
    """
    prev_excel, prev_weeks, prev_week_cols = prev_data
    new_excel,  new_weeks,  new_week_cols  = new_data

    duplicates = set(prev_weeks) & set(new_weeks)
    if duplicates:
        dup_list = ", ".join(f"CW{w:02d}" for w in sorted(duplicates))
        raise ValueError(
            f"The following week(s) already exist in your previous waterfall "
            f"and cannot be added again: **{dup_list}**.\n\n"
            f"Please remove the duplicate file(s) and try again."
        )

    return (
        prev_excel + new_excel,
        prev_weeks + new_weeks,
        prev_week_cols | new_week_cols,
    )


# ── Shared Excel formatting ───────────────────────────────────────────────────
def blank_pre_snapshot_weeks(waterfall: pd.DataFrame, row_file_indices: list,
                             snapshot_weeks: list, all_weeks: list) -> None:
    """Zero-out week columns that predate each row's snapshot week (in-place)."""
    for row_idx, file_idx in enumerate(row_file_indices):
        if file_idx is None:
            continue
        snapshot_week = next(
            (w for w in all_weeks
             if int(w.split("-")[0][1:]) == snapshot_weeks[file_idx]),
            None,
        )
        if snapshot_week is None:
            continue
        for col_pos in range(all_weeks.index(snapshot_week)):
            waterfall.at[row_idx, all_weeks[col_pos]] = ""


def apply_excel_formatting(ws, header: list, col_name_to_idx: dict,
                           row_file_indices: list, group_index: list,
                           id_cols: set, snapshot_weeks: list,
                           all_weeks: list, var_col_data: dict) -> None:
    """Apply all cell-level formatting to an openpyxl worksheet."""
    # Header row
    for col_idx, _ in enumerate(header, start=1):
        cell           = ws.cell(row=1, column=col_idx)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center_align

    # Data rows
    for row_idx, file_idx in enumerate(row_file_indices):
        excel_row = row_idx + 2

        if file_idx is None:
            for col_idx in range(1, len(header) + 1):
                ws.cell(row=excel_row, column=col_idx).fill = sep_fill
            continue

        is_alt = (group_index[row_idx] % 2 == 1)

        for col_idx, col_name in enumerate(header, start=1):
            cell = ws.cell(row=excel_row, column=col_idx)
            if col_name in id_cols:
                cell.fill      = solid(ID_COL_BG_ALT if is_alt else ID_COL_BG)
                cell.alignment = left_align
                cell.font      = Font(name="Arial", size=10)
            elif col_name in VAR_COLS:
                cell.fill      = solid(VAR_COL_BG_ALT if is_alt else VAR_COL_BG)
                cell.alignment = center_align
                cell.font      = Font(name="Arial", size=10)
            elif col_name == "SnapshotWeek":
                cell.fill      = solid(YELLOW)
                cell.alignment = center_align
                cell.font      = Font(name="Arial", bold=True, size=10)
            else:
                cell.fill      = solid(WEEK_COL_BG_ALT if is_alt else WEEK_COL_BG)
                cell.alignment = center_align
                cell.font      = Font(name="Arial", size=10)

        if file_idx < len(snapshot_weeks):
            target_week_col = next(
                (w for w in all_weeks
                 if int(w.split("-")[0][1:]) == snapshot_weeks[file_idx]),
                None,
            )
            if target_week_col and target_week_col in col_name_to_idx:
                ws.cell(row=excel_row,
                        column=col_name_to_idx[target_week_col]).fill = yellow_fill

        for col_name, col_data in var_col_data.items():
            col_idx = col_name_to_idx.get(col_name)
            if not col_idx:
                continue
            raw_val = col_data[row_idx]
            cell    = ws.cell(row=excel_row, column=col_idx)
            if isinstance(raw_val, float):
                cell.value         = raw_val
                cell.number_format = "0.0%"
                if abs(raw_val) >= RED_THRESHOLD[col_name]:
                    cell.fill = red_fill
                    cell.font = white_font
